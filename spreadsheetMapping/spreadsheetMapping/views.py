from django.shortcuts import render
from django.apps import apps
from django.db.utils import IntegrityError
import openpyxl as px
import re

MODELS = [
    model.__name__ 
    for model in apps.get_models()
]

RE_LETTERS = re.compile(r"([A-Z]+)")

def upload_spreadsheet(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        return ingest_spreadsheet(request, uploaded_file)
    return render(request, 'upload.html')

def ingest_spreadsheet(request, uploaded_file):

    def callback(Model, sheet):
        column_header_mapping = {
            field.help_text.lower(): field.name.lower()
            for field in Model._meta.fields
        }

        # Remove the id field
        del column_header_mapping['']

        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for cell in row:
                letters = RE_LETTERS.match(cell.coordinate).group(1)
                row_data[letters.lower()] = cell.value

            # Shuffle the data into the correct order as it appears in the model
            shuffled_row_data = {
                column_header_mapping[key]: row_data[key]
                for key in column_header_mapping.keys()
            }

            try:
                instance = Model(**shuffled_row_data)
                instance.save()
            except IntegrityError as e:
                print(f"Error saving {shuffled_row_data} to {Model.__name__} due to {e}")

    workbook = px.load_workbook(uploaded_file)
    considering = list(filter(lambda x: x in MODELS, workbook.sheetnames))
    
    for name in considering:
        Model = apps.get_model(app_label="spreadsheetMapping", model_name=name)
        sheet = workbook[name]
        callback(Model, sheet)

    return render(request, 'success.html')
